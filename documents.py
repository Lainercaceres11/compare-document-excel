"""
comparar_tiquetes_gui.py  —  Herramienta de cruce COMFACHOCO
═══════════════════════════════════════════════════════════════
Compatible con CUALQUIER variante del archivo MES y TIQUETES:

  ARCHIVO MES
    · Encabezado en fila 1 (FEBRERO_2026) o fila 2 (MARZO_2026) — auto-detectado
    · Columna TIQUETE con nombre "NUMERO DEL TIQUETE" (FEBRERO) o sin nombre (MARZO)
    · Nombres de columnas en mayúsculas o mixtos ("Fecha","Hora","Total")

  ARCHIVO TIQUETES
    · .xls (FEBRERO) o .xlsx (MARZO)
    · Encabezado auto-detectado buscando la celda "TIQUETE"
    · Columnas requeridas: TIQUETE, CEDULA PASAJERO, FEC.SALIDA, NRO ORDEN CREDITO

ESTRATEGIA DE COMPARACIÓN — cuatro niveles de prioridad:

    Nivel 1 — Exacto:    orden_tiq == orden_mes
    Nivel 2 — Prefijo:   orden_tiq.startswith(orden_mes)
              (sistemas que añaden dígitos al final,
               ej. MES=2026017179 ↔ TIQUETES=20260171799)
    Nivel 3 — Aproximado protegido:
              |int(orden_tiq) − int(orden_mes)| ≤ TOLERANCIA_ORDEN
              SOLO si orden_tiq NO tiene match exacto/prefijo con ninguna
              OTRA orden del mismo pasajero en el MES.
    Nivel 4 — Sin número de orden (NRO ORDEN CREDITO = 0 o vacío):
              Asignación por cédula sola, prioridad mínima.

  Candidatos ordenados por nivel, luego cronológicamente (FEC.SALIDA).
  Cada tiquete se asigna UNA SOLA VEZ.
"""

import subprocess, os, shutil, threading
import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox


TOLERANCIA_ORDEN = 20


# ══════════════════════════════════════════════════════════════════
#  DETECCIÓN AUTOMÁTICA DE ESTRUCTURA
# ══════════════════════════════════════════════════════════════════

def detectar_fila(ws, texto, max_filas=10):
    """
    Fila (1-indexada) donde aparece 'texto'. Lanza RuntimeError si no encuentra.
    """
    for row in range(1, max_filas + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if v is not None and str(v).strip() == texto:
                return row
    raise RuntimeError(
        f"No se encontró '{texto}' en las primeras {max_filas} filas.\n"
        "Verifique que seleccionó el archivo correcto."
    )


def construir_col_map(ws, header_row):
    """Dict {nombre_columna: número_columna} desde la fila de encabezado."""
    return {
        ws.cell(header_row, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value is not None
    }


def buscar_col(col_map, *nombres):
    """
    Número de columna por nombre, insensible a mayúsculas/espacios.
    Devuelve None si no encuentra ninguno.
    """
    upper = {k.strip().upper(): v for k, v in col_map.items() if k is not None}
    for n in nombres:
        key = n.strip().upper()
        if key in upper:
            return upper[key]
    return None


def detectar_col_tiq(ws, col_map, header_row):
    """
    Columna donde se escribirá el número de tiquete:
    · FEBRERO: columna con nombre "NUMERO DEL TIQUETE"
    · MARZO:   primera columna SIN nombre antes de "Nro orden de compra"
    """
    col = buscar_col(col_map, "NUMERO DEL TIQUETE", "NUMERO TIQUETE", "NRO TIQUETE")
    if col is not None:
        return col
    # Fallback: primera columna vacía antes de la columna de orden
    col_orden = buscar_col(col_map, "Nro orden de compra") or ws.max_column
    for c in range(1, col_orden):
        if ws.cell(header_row, c).value is None:
            return c
    return None


# ══════════════════════════════════════════════════════════════════
#  LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════

def buscar_libreoffice():
    import shutil as sh
    found = sh.which("libreoffice") or sh.which("soffice")
    if found:
        return found
    for r in [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]:
        if os.path.isfile(r):
            return r
    return None


def convertir_xls(ruta_xls, ejecutable, log_fn):
    """
    Convierte .xls → .xlsx con LibreOffice.
    Invalida la caché si el .xls es más reciente.
    Busca el resultado en múltiples directorios.
    """
    carpeta = os.path.dirname(os.path.abspath(ruta_xls))
    base    = os.path.splitext(os.path.basename(ruta_xls))[0]
    destino = os.path.join(carpeta, base + "_converted.xlsx")

    if os.path.exists(destino):
        if os.path.getmtime(ruta_xls) > os.path.getmtime(destino):
            log_fn("INFO", "  Caché obsoleta, reconvirtiendo .xls...")
            os.remove(destino)
        else:
            log_fn("INFO", "  Usando conversión en caché...")
            return destino

    cwd = os.getcwd()
    result = subprocess.run(
        [ejecutable, "--headless", "--convert-to", "xlsx",
         os.path.abspath(ruta_xls), "--outdir", carpeta],
        capture_output=True, text=True, cwd=carpeta,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice no pudo convertir el archivo.\n{result.stderr}")

    nombre_xlsx = base + ".xlsx"
    posibles = [
        os.path.join(carpeta, nombre_xlsx),
        os.path.join(cwd, nombre_xlsx),
        os.path.join("/", nombre_xlsx),
        os.path.join(os.path.expanduser("~"), nombre_xlsx),
    ]
    encontrado = next((p for p in posibles if os.path.exists(p)), None)
    if encontrado is None:
        raise RuntimeError(
            f"LibreOffice terminó pero no se encontró '{nombre_xlsx}'.\n"
            f"Buscado en: {posibles}"
        )
    if encontrado != destino:
        shutil.move(encontrado, destino)
    return destino


def leer_mes(ruta):
    """
    Lee el archivo MES. Detecta automáticamente la fila del encabezado
    buscando 'Nro orden de compra'.
    Devuelve (df, header_excel_row).
    """
    wb = load_workbook(ruta, read_only=True)
    ws = wb.active
    hrow = detectar_fila(ws, "Nro orden de compra")
    wb.close()
    df = pd.read_excel(ruta, header=hrow - 1)
    return df, hrow


def leer_tiquetes(ruta, log_fn):
    """
    Lee TIQUETES EXPEDIDOS (.xls o .xlsx).
    Detecta automáticamente la fila del encabezado buscando 'TIQUETE'.
    """
    ext = os.path.splitext(ruta)[1].lower()

    if ext == ".xlsx":
        wb = load_workbook(ruta, read_only=True)
        ws = wb.active
        hrow = detectar_fila(ws, "TIQUETE")
        wb.close()
        log_fn("INFO", f"  Encabezado TIQUETES en fila {hrow}")
        df = pd.read_excel(ruta, header=hrow - 1)
        df.columns = df.columns.str.strip()
        return df

    # .xls → convertir con LibreOffice
    lo = buscar_libreoffice()
    if lo:
        log_fn("INFO", "  Convirtiendo .xls con LibreOffice...")
        ruta_conv = convertir_xls(ruta, lo, log_fn)
        wb = load_workbook(ruta_conv, read_only=True)
        ws = wb.active
        hrow = detectar_fila(ws, "TIQUETE")
        wb.close()
        log_fn("INFO", f"  Encabezado TIQUETES en fila {hrow}")
        df = pd.read_excel(ruta_conv, header=hrow - 1)
        df.columns = df.columns.str.strip()
        if df.empty or "CEDULA PASAJERO" not in df.columns:
            raise RuntimeError(
                "El archivo de tiquetes no tiene la estructura esperada.\n"
                f"Columnas: {list(df.columns)}"
            )
        return df

    # Fallback: xlrd
    log_fn("INFO", "  Usando xlrd para leer .xls...")
    try:
        import xlrd  # noqa
    except ImportError:
        raise ImportError("Instala xlrd: pip install xlrd  o instala LibreOffice.")
    df_raw = pd.read_excel(ruta, header=None, engine="xlrd")
    hrow_idx = next(
        i for i, row in df_raw.iterrows()
        if any(str(v).strip().upper() == "TIQUETE" for v in row)
    )
    df_raw.columns = df_raw.iloc[hrow_idx].str.strip()
    return df_raw.iloc[hrow_idx + 1:].reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════
#  UTILIDADES DE DATOS
# ══════════════════════════════════════════════════════════════════

def to_str_orden(v):
    try:
        return str(int(float(v))) if v is not None and not pd.isna(v) else ""
    except (ValueError, TypeError):
        return str(v).strip() if v else ""


def to_str_doc(v):
    try:
        return str(int(float(v))) if v is not None and not pd.isna(v) else ""
    except (ValueError, TypeError):
        return str(v).strip() if v else ""


def separar_fecha_hora(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None, None
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y"), valor.strftime("%H:%M:%S")
    try:
        dt = datetime.strptime(str(valor).strip(), "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M:%S")
    except ValueError:
        return str(valor), None


def exacto_o_prefijo(o_tiq, o_mes):
    return bool(o_tiq) and bool(o_mes) and (o_tiq == o_mes or o_tiq.startswith(o_mes))


def get_candidatos(o_mes, d_mes, tiq_index, otras_ordenes):
    """
    Candidatos para la fila (o_mes, d_mes), ordenados por prioridad:
      0   — Exacto/prefijo con esta orden
      1…N — Aproximado protegido (no pertenece a otra orden del mismo pasajero)
      999 — Sin número de orden (por cédula sola)
    """
    cands = []
    for e in tiq_index.get(d_mes, []):
        o_tiq = e["orden_str"]
        if exacto_o_prefijo(o_tiq, o_mes):
            cands.append((0, e))
        elif o_tiq in ("", "0"):
            cands.append((999, e))
        else:
            # Aproximado: solo si ninguna OTRA orden del pasajero lo reclama
            if not any(exacto_o_prefijo(o_tiq, o) for o in otras_ordenes):
                try:
                    diff = abs(int(o_tiq) - int(o_mes))
                    if diff <= TOLERANCIA_ORDEN:
                        cands.append((diff, e))
                except (ValueError, TypeError):
                    pass
    cands.sort(key=lambda x: (x[0], str(x[1].get("raw_fec", "") or "")))
    return [e for _, e in cands]


# ══════════════════════════════════════════════════════════════════
#  NÚCLEO DEL PROCESO
# ══════════════════════════════════════════════════════════════════

_wb_orig_ws = None


def procesar(ruta_mes, ruta_tiquetes, log_fn, progress_fn):
    global _wb_orig_ws

    # ── Leer MES ───────────────────────────────────────────────────
    progress_fn(5)
    log_fn("INFO", "Leyendo archivo MES...")
    df_mes, hrow_mes = leer_mes(ruta_mes)
    wb_orig = load_workbook(ruta_mes)
    _wb_orig_ws = wb_orig.active
    log_fn("INFO", f"  Encabezado en fila {hrow_mes} | {len(df_mes)} filas de datos")

    # ── Leer TIQUETES ──────────────────────────────────────────────
    progress_fn(15)
    log_fn("INFO", "Leyendo archivo TIQUETES EXPEDIDOS...")
    df_tiq = leer_tiquetes(ruta_tiquetes, log_fn)
    log_fn("INFO", f"  {len(df_tiq)} tiquetes cargados")

    # Validar columnas requeridas
    cols_tiq = set(df_tiq.columns.tolist())
    for req in ("TIQUETE", "CEDULA PASAJERO", "FEC.SALIDA", "NRO ORDEN CREDITO"):
        if req not in cols_tiq:
            raise RuntimeError(
                f"Columna '{req}' no encontrada en TIQUETES.\n"
                f"Columnas disponibles: {sorted(cols_tiq)}"
            )

    # ── Detectar columnas del MES ──────────────────────────────────
    col_map_mes = construir_col_map(_wb_orig_ws, hrow_mes)
    upper_mes   = {k.strip().upper(): v for k, v in col_map_mes.items()}

    col_cant_r  = upper_mes.get("CANTIDAD DE PASAJES", 15)
    col_tar_r   = upper_mes.get("TARIFA", 16)
    col_tiq_r   = detectar_col_tiq(_wb_orig_ws, col_map_mes, hrow_mes)
    col_fec_r   = upper_mes.get("FECHA", 4)
    col_hora_r  = upper_mes.get("HORA", 5)
    col_tot_r   = upper_mes.get("TOTAL", 17)

    if col_tiq_r is None:
        raise RuntimeError(
            "No se detectó la columna de TIQUETE en el archivo MES.\n"
            f"Columnas disponibles: {sorted(col_map_mes.keys())}"
        )
    log_fn("INFO",
           f"  Columnas detectadas: TIQUETE=col{col_tiq_r}, "
           f"FECHA=col{col_fec_r}, HORA=col{col_hora_r}")

    # ── Paso 1: índice de tiquetes ─────────────────────────────────
    progress_fn(28)
    log_fn("INFO", "Paso 1 — Indexando tiquetes por cédula...")

    tiq_index = {}
    sin_cedula = 0
    for _, row in df_tiq.iterrows():
        o_str = to_str_orden(row.get("NRO ORDEN CREDITO"))
        d_str = to_str_doc(row.get("CEDULA PASAJERO"))
        if not d_str:
            sin_cedula += 1
            continue
        fecha, hora = separar_fecha_hora(row.get("FEC.SALIDA"))
        tiq_index.setdefault(d_str, []).append({
            "orden_str": o_str,
            "tiquete":   str(row["TIQUETE"]).strip(),
            "fecha":     fecha,
            "hora":      hora,
            "raw_fec":   row.get("FEC.SALIDA"),
        })

    log_fn("INFO", f"  Cédulas únicas en TIQUETES: {len(tiq_index)}")
    if sin_cedula:
        log_fn("INFO", f"  Filas sin cédula: {sin_cedula}")

    # ── Paso 2: mapa cédula → órdenes en MES ──────────────────────
    progress_fn(35)
    log_fn("INFO", "Paso 2 — Mapeando órdenes por pasajero...")

    ced_a_ordenes = {}
    for df_idx in df_mes.index:
        o = to_str_orden(df_mes.loc[df_idx, "Nro orden de compra"])
        d = to_str_doc(df_mes.loc[df_idx, "Número de Documento"])
        if d and o:
            ced_a_ordenes.setdefault(d, set()).add(o)

    # ── Paso 3: planificar asignación ─────────────────────────────
    progress_fn(40)
    log_fn("INFO", "Paso 3 — Cruzando tiquetes con autorizaciones...")

    plan = []
    filas_ok = filas_no = filas_aprox = 0

    for df_idx in df_mes.index:
        excel_row_orig = df_idx + hrow_mes + 1   # fila Excel en el workbook original
        o_mes = to_str_orden(df_mes.loc[df_idx, "Nro orden de compra"])
        d_mes = to_str_doc(df_mes.loc[df_idx, "Número de Documento"])

        val_cant = _wb_orig_ws.cell(excel_row_orig, col_cant_r).value
        try:
            cantidad = int(val_cant) if val_cant and int(val_cant) > 0 else 1
        except (TypeError, ValueError):
            cantidad = 1

        otras = ced_a_ordenes.get(d_mes, set()) - {o_mes}
        candidatos = get_candidatos(o_mes, d_mes, tiq_index, otras)

        if candidatos:
            filas_ok += 1
            if any(
                not exacto_o_prefijo(e["orden_str"], o_mes) and e["orden_str"] not in ("", "0")
                for e in candidatos
            ):
                filas_aprox += 1
        else:
            filas_no += 1

        plan.append((excel_row_orig, cantidad, candidatos))

    log_fn("OK",   f"Filas con tiquete encontrado: {filas_ok}")
    if filas_aprox:
        log_fn("INFO", f"  De los cuales por match aproximado: {filas_aprox}")
    log_fn("WARN", f"Filas sin tiquete (no disponibles): {filas_no}")

    # ── Preparar archivo de salida ────────────────────────────────
    progress_fn(55)
    log_fn("INFO", "Copiando formato del archivo MES original...")

    ruta_salida = os.path.join(
        os.path.dirname(os.path.abspath(ruta_mes)),
        os.path.splitext(os.path.basename(ruta_mes))[0] + "_ACTUALIZADO.xlsx"
    )
    shutil.copy2(ruta_mes, ruta_salida)

    wb = load_workbook(ruta_salida)
    ws = wb.active

    n_cols   = ws.max_column
    amarillo = PatternFill("solid", fgColor="FFFF00")

    # ── Expandir filas y escribir ──────────────────────────────────
    progress_fn(68)
    log_fn("INFO", "Expandiendo filas y escribiendo datos...")

    primera_fila_datos = hrow_mes + 1
    ws.delete_rows(primera_fila_datos, ws.max_row - hrow_mes)

    tiquetes_usados  = set()
    filas_expandidas = 0
    write_row        = primera_fila_datos

    cols_limpiar = (col_tiq_r, col_fec_r, col_hora_r)

    for (orig_row, cantidad, candidatos) in plan:
        disponibles = [e for e in candidatos if e["tiquete"] not in tiquetes_usados]

        for _rep in range(cantidad):
            ws.append([None] * n_cols)

            for c in range(1, n_cols + 1):
                src = _wb_orig_ws.cell(orig_row, c)
                dst = ws.cell(write_row, c)
                if src.has_style:
                    dst.font          = copy(src.font)
                    dst.fill          = copy(src.fill)
                    dst.border        = copy(src.border)
                    dst.alignment     = copy(src.alignment)
                    dst.number_format = src.number_format
                if c == col_cant_r:
                    dst.value = 1
                elif c == col_tot_r:
                    dst.value = _wb_orig_ws.cell(orig_row, col_tar_r).value or 0
                elif c in cols_limpiar:
                    dst.value = None   # se limpia; si hay tiquete se sobreescribe abajo
                else:
                    dst.value = src.value

            if disponibles:
                entry = disponibles.pop(0)
                tiquetes_usados.add(entry["tiquete"])
                ws.cell(write_row, col_tiq_r).value = entry["tiquete"]
                ws.cell(write_row, col_fec_r).value  = entry["fecha"]
                ws.cell(write_row, col_hora_r).value = entry["hora"]
                for col in cols_limpiar:
                    ws.cell(write_row, col).fill = amarillo

            if cantidad > 1:
                filas_expandidas += 1
            write_row += 1

    progress_fn(90)
    log_fn("INFO", "Guardando archivo actualizado...")
    wb.save(ruta_salida)
    progress_fn(100)

    total_filas   = write_row - primera_fila_datos
    tiq_asignados = len(tiquetes_usados)
    log_fn("OK", f"Tiquetes asignados: {tiq_asignados}")
    log_fn("OK", f"Filas expandidas (pasajes > 1): {filas_expandidas}")
    log_fn("OK", f"Total filas en archivo final: {total_filas}")

    return ruta_salida, filas_ok, filas_no, tiq_asignados, total_filas


# ══════════════════════════════════════════════════════════════════
#  PALETA Y TIPOGRAFÍA
# ══════════════════════════════════════════════════════════════════

C = {
    "bg":            "#F7F6F3",
    "surface":       "#FFFFFF",
    "surface2":      "#F0EEE9",
    "border":        "#E2DED6",
    "border_focus":  "#1A1A1A",
    "text":          "#1A1A1A",
    "text2":         "#6B6860",
    "text3":         "#A09D97",
    "accent":        "#1A1A1A",
    "accent_hover":  "#333333",
    "ok":            "#1B7A4A",
    "ok_bg":         "#EBF7F1",
    "warn":          "#9A6500",
    "warn_bg":       "#FFF8E6",
    "err":           "#C0392B",
    "err_bg":        "#FDECEA",
    "info":          "#1A4A7A",
    "info_bg":       "#EBF2FB",
    "progress_bg":   "#E8E6E1",
    "progress_fill": "#1A1A1A",
}

FT = {
    "title": ("Georgia", 18, "bold"),
    "sub":   ("Georgia", 10, "italic"),
    "label": ("Helvetica", 9, "bold"),
    "body":  ("Helvetica", 10),
    "small": ("Helvetica", 8),
    "log":   ("Courier", 9),
    "btn":   ("Helvetica", 11, "bold"),
    "stat":  ("Georgia", 22, "bold"),
    "statl": ("Helvetica", 8),
}


# ══════════════════════════════════════════════════════════════════
#  COMPONENTES UI
# ══════════════════════════════════════════════════════════════════

class FileCard(tk.Frame):
    def __init__(self, parent, number, label, extensions, **kw):
        super().__init__(parent, bg=C["surface"],
                         highlightbackground=C["border"], highlightthickness=1, **kw)
        self.var = tk.StringVar()

        nf = tk.Frame(self, bg=C["accent"], width=32, height=32)
        nf.pack_propagate(False)
        nf.pack(side="left", padx=(16, 12), pady=16)
        tk.Label(nf, text=str(number), font=("Helvetica", 11, "bold"),
                 fg=C["surface"], bg=C["accent"]).place(relx=0.5, rely=0.5, anchor="center")

        center = tk.Frame(self, bg=C["surface"])
        center.pack(side="left", fill="x", expand=True, pady=12)
        tk.Label(center, text=label.upper(), font=FT["label"],
                 fg=C["text2"], bg=C["surface"], anchor="w").pack(fill="x")
        self.path_lbl = tk.Label(center, text="Haz clic para seleccionar archivo...",
                                 font=FT["body"], fg=C["text3"], bg=C["surface"], anchor="w")
        self.path_lbl.pack(fill="x")

        tk.Button(self, text="Examinar", font=FT["small"],
                  fg=C["text"], bg=C["surface2"],
                  activebackground=C["border"], activeforeground=C["text"],
                  relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
                  command=lambda: self._browse(extensions)
                  ).pack(side="right", padx=16, pady=16)

        self.dot = tk.Label(self, text="*", font=("Helvetica", 14, "bold"),
                            fg=C["border"], bg=C["surface"])
        self.dot.place(relx=1.0, rely=0.0, x=-10, y=6, anchor="ne")

        for w in [self, center, self.path_lbl]:
            w.bind("<Enter>", self._on_enter)
            w.bind("<Leave>", self._on_leave)
            w.bind("<Button-1>", lambda e, ext=extensions: self._browse(ext))

    def _browse(self, extensions):
        path = filedialog.askopenfilename(filetypes=[("Excel", extensions)])
        if path:
            self.var.set(path)
            self.path_lbl.configure(text=os.path.basename(path), fg=C["text"])
            self.dot.configure(fg=C["ok"])
            self.configure(highlightbackground=C["ok"])

    def _on_enter(self, e):
        if not self.var.get():
            self.configure(highlightbackground=C["border_focus"])

    def _on_leave(self, e):
        if not self.var.get():
            self.configure(highlightbackground=C["border"])

    def get(self):
        return self.var.get().strip()


class StatCard(tk.Frame):
    def __init__(self, parent, label, color_key="ok", **kw):
        bg, fg = C[color_key + "_bg"], C[color_key]
        super().__init__(parent, bg=bg,
                         highlightbackground=C["border"], highlightthickness=1, **kw)
        self._var = tk.StringVar(value="--")
        tk.Label(self, textvariable=self._var, font=FT["stat"], fg=fg, bg=bg).pack(pady=(14, 2))
        tk.Label(self, text=label.upper(), font=FT["statl"], fg=fg, bg=bg,
                 wraplength=110, justify="center").pack(pady=(0, 14))

    def set(self, val):
        self._var.set(str(val))


class LogLine(tk.Frame):
    _ICONS  = {"OK": "+", "WARN": "!", "ERR": "X", "INFO": "."}
    _COLORS = {
        "OK":   (C["ok"],    C["ok_bg"]),
        "WARN": (C["warn"],  C["warn_bg"]),
        "ERR":  (C["err"],   C["err_bg"]),
        "INFO": (C["text2"], C["surface"]),
    }

    def __init__(self, parent, level, message, **kw):
        fg, bg = self._COLORS.get(level, (C["text2"], C["surface"]))
        super().__init__(parent, bg=bg, **kw)
        tk.Label(self, text=self._ICONS.get(level, "."),
                 font=("Helvetica", 9, "bold"), fg=fg, bg=bg, width=2
                 ).pack(side="left", padx=(8, 4), pady=3)
        tk.Label(self, text=message, font=FT["log"],
                 fg=fg if level != "INFO" else C["text2"],
                 bg=bg, anchor="w"
                 ).pack(side="left", fill="x", pady=3, padx=(0, 8))


class AnimatedButton(tk.Button):
    def __init__(self, parent, **kw):
        super().__init__(parent,
                         bg=C["accent"], fg=C["surface"],
                         activebackground=C["accent_hover"], activeforeground=C["surface"],
                         relief="flat", bd=0, cursor="hand2", font=FT["btn"], pady=14, **kw)
        self.bind("<Enter>", lambda e: self.configure(bg=C["accent_hover"]))
        self.bind("<Leave>", lambda e: self.configure(bg=C["accent"]))


class ProgressBar(tk.Canvas):
    def __init__(self, parent, **kw):
        super().__init__(parent, height=4, bg=C["progress_bg"], highlightthickness=0, **kw)
        self._pct = self._target = 0.0
        self._bar = self.create_rectangle(0, 0, 0, 4, fill=C["progress_fill"], width=0)
        self.bind("<Configure>", self._draw)

    def _draw(self, e=None):
        self.coords(self._bar, 0, 0, int(self.winfo_width() * self._pct / 100), 4)

    def set_target(self, pct):
        self._target = float(pct)
        self._step()

    def _step(self):
        if abs(self._pct - self._target) < 0.5:
            self._pct = self._target
            self._draw()
            return
        self._pct += (self._target - self._pct) * 0.15
        self._draw()
        self.after(16, self._step)

    def reset(self):
        self._pct = self._target = 0.0
        self._draw()


# ══════════════════════════════════════════════════════════════════
#  VENTANA PRINCIPAL
# ══════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cruce COMFACHOCO — Tiquetes vs Autorizaciones")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self._build()

    def _build(self):
        outer = tk.Frame(self, bg=C["bg"])
        outer.pack(fill="both", expand=True, padx=32, pady=28)

        hdr = tk.Frame(outer, bg=C["bg"])
        hdr.pack(fill="x", pady=(0, 6))
        tk.Label(hdr, text="Cruce", font=FT["title"],
                 fg=C["text"], bg=C["bg"]).pack(side="left")
        tk.Label(hdr, text=" COMFACHOCO", font=("Georgia", 18),
                 fg=C["text2"], bg=C["bg"]).pack(side="left")
        tk.Label(hdr, text="Tiquetes vs Autorizaciones",
                 font=FT["sub"], fg=C["text3"], bg=C["bg"]
                 ).pack(side="right", anchor="s", pady=4)

        tk.Label(outer,
                 text="Compara cedula y orden entre archivos y completa "
                      "tiquete, fecha y hora. Compatible con FEBRERO y MARZO 2026.",
                 font=FT["small"], fg=C["text3"], bg=C["bg"],
                 anchor="w", justify="left", wraplength=480).pack(fill="x", pady=(0, 16))

        tk.Frame(outer, bg=C["border"], height=1).pack(fill="x", pady=(0, 20))

        tk.Label(outer, text="ARCHIVOS DE ENTRADA", font=FT["label"],
                 fg=C["text3"], bg=C["bg"], anchor="w").pack(fill="x", pady=(0, 8))

        self.card_mes = FileCard(outer, 1,
                                 "Archivo MES - Autorizaciones COMFACHOCO",
                                 "*.xlsx *.xls")
        self.card_mes.pack(fill="x", pady=(0, 8))

        self.card_tiq = FileCard(outer, 2,
                                 "Archivo TIQUETES EXPEDIDOS",
                                 "*.xlsx *.xls")
        self.card_tiq.pack(fill="x")

        nota = tk.Frame(outer, bg=C["info_bg"],
                        highlightbackground=C["border"], highlightthickness=1)
        nota.pack(fill="x", pady=(10, 0))
        tk.Label(nota,
                 text="El archivo actualizado se genera automaticamente "
                      "en la misma carpeta que el archivo MES.",
                 font=FT["small"], fg=C["info"], bg=C["info_bg"],
                 anchor="w", pady=8, padx=12).pack(fill="x")

        tk.Frame(outer, bg=C["border"], height=1).pack(fill="x", pady=(20, 16))

        self.progress = ProgressBar(outer, width=480)
        self.progress.pack(fill="x", pady=(0, 6))
        self.status_var = tk.StringVar(value="Listo para procesar")
        tk.Label(outer, textvariable=self.status_var,
                 font=FT["small"], fg=C["text3"], bg=C["bg"], anchor="w").pack(fill="x")

        self.btn = AnimatedButton(outer, text="Procesar archivos", command=self._run)
        self.btn.pack(fill="x", pady=(14, 0))

        tk.Frame(outer, bg=C["border"], height=1).pack(fill="x", pady=(20, 16))
        tk.Label(outer, text="RESULTADOS", font=FT["label"],
                 fg=C["text3"], bg=C["bg"], anchor="w").pack(fill="x", pady=(0, 10))

        sf = tk.Frame(outer, bg=C["bg"])
        sf.pack(fill="x")
        self.s_ok    = StatCard(sf, "Filas con tiquete",  "ok")
        self.s_no    = StatCard(sf, "Sin tiquete",        "warn")
        self.s_tiq   = StatCard(sf, "Tiquetes asignados", "ok")
        self.s_total = StatCard(sf, "Total filas",        "info")
        for i, s in enumerate([self.s_ok, self.s_no, self.s_tiq, self.s_total]):
            s.grid(row=0, column=i, padx=(0, 8) if i < 3 else 0, sticky="nsew")
            sf.columnconfigure(i, weight=1)

        tk.Frame(outer, bg=C["border"], height=1).pack(fill="x", pady=(20, 16))
        tk.Label(outer, text="REGISTRO DE ACTIVIDAD", font=FT["label"],
                 fg=C["text3"], bg=C["bg"], anchor="w").pack(fill="x", pady=(0, 8))

        log_wrap = tk.Frame(outer, bg=C["surface"],
                            highlightbackground=C["border"], highlightthickness=1)
        log_wrap.pack(fill="both")
        self.log_inner = tk.Frame(log_wrap, bg=C["surface"])
        self.log_inner.pack(fill="both", expand=True)
        self._placeholder = tk.Label(
            self.log_inner,
            text="El registro aparecera aqui una vez iniciado el proceso.",
            font=FT["small"], fg=C["text3"], bg=C["surface"], pady=20)
        self._placeholder.pack()
        self._log_lines = []

    def _log(self, level, msg):
        if self._placeholder.winfo_ismapped():
            self._placeholder.pack_forget()
        line = LogLine(self.log_inner, level, msg)
        line.pack(fill="x")
        self._log_lines.append(line)
        self.log_inner.update_idletasks()

    def _clear_log(self):
        for w in self._log_lines:
            w.destroy()
        self._log_lines = []
        self._placeholder.pack()

    def _restaurar_btn(self):
        self.btn.configure(state="normal", text="Procesar archivos")
        self.progress.set_target(0)

    def _run(self):
        mes = self.card_mes.get()
        tiq = self.card_tiq.get()
        if not mes or not tiq:
            messagebox.showwarning("Archivos requeridos",
                                   "Selecciona ambos archivos antes de continuar.")
            return

        self.btn.configure(state="disabled", text="Procesando...")
        self.progress.reset()
        self._clear_log()
        for s in [self.s_ok, self.s_no, self.s_tiq, self.s_total]:
            s.set("--")

        def log_fn(level, msg):
            self.after(0, lambda lv=level, m=msg: self._log(lv, m))

        def progress_fn(pct):
            self.after(0, lambda p=pct: self.progress.set_target(p))
            self.after(0, lambda p=pct: self.status_var.set(f"Procesando... {p}%"))

        def worker():
            try:
                ruta, ok, no, tiq_n, total = procesar(mes, tiq, log_fn, progress_fn)
                ruta_c, ok_c, no_c, tiq_c, total_c = ruta, ok, no, tiq_n, total
                def done():
                    self.s_ok.set(ok_c)
                    self.s_no.set(no_c)
                    self.s_tiq.set(tiq_c)
                    self.s_total.set(total_c)
                    self.status_var.set(f"Completado — {os.path.basename(ruta_c)}")
                    self._restaurar_btn()
                    messagebox.showinfo(
                        "Proceso completado",
                        f"Archivo generado exitosamente.\n\nUbicacion:\n{ruta_c}"
                    )
                self.after(0, done)
            except Exception as exc:
                err_msg = str(exc)
                def on_err(msg=err_msg):
                    self._log("ERR", msg)
                    self.status_var.set("Error durante el proceso")
                    self._restaurar_btn()
                    messagebox.showerror("Error", msg)
                self.after(0, on_err)

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()